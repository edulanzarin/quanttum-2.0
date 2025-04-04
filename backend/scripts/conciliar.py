import pandas as pd
import tkinter as tk
import json
import os
import sys
from tkinter import filedialog
import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import load_workbook

CAMINHO_JSON = os.path.join(os.path.dirname(
    __file__), '..', 'database', 'serviceAccountKey.json')

# Configuração do Firebase
cred = credentials.Certificate(CAMINHO_JSON)
firebase_admin.initialize_app(cred)
db = firestore.client()


def salvar_dados(nome_sugerido):
    try:
        root = tk.Tk()
        root.withdraw()  # Não exibe a janela principal do Tkinter
        root.attributes("-topmost", 1)  # Mantém a janela de diálogo no topo
        # Remove o topo após 100ms para normalizar
        root.after(100, lambda: root.attributes("-topmost", 0))

        caminho_arquivo = filedialog.asksaveasfilename(
            initialfile=nome_sugerido,
            defaultextension=".xlsx",  # Alterado para .xlsx
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]  # Atualizado para Excel
        )

        root.destroy()  # Fecha a janela do Tkinter corretamente
        return caminho_arquivo if caminho_arquivo else None
    except Exception:
        return None
    
def carregar_planilha(caminho_arquivo):
    """Carrega uma planilha CSV ou XLSX."""
    try:
        if caminho_arquivo is None:
            raise ValueError("O caminho do arquivo não pode ser None.")

        if caminho_arquivo.endswith('.csv'):
            df = pd.read_csv(caminho_arquivo, delimiter=';', encoding='latin1')
        elif caminho_arquivo.endswith('.xlsx'):
            # Carrega o arquivo Excel com openpyxl para verificar a formatação
            wb = load_workbook(caminho_arquivo)
            ws = wb.active

            # Converte a planilha para DataFrame
            df = pd.DataFrame(ws.values)
            df.columns = df.iloc[0]  # Define a primeira linha como cabeçalho
            df = df[1:]  # Remove a primeira linha

            # Verifica a formatação das células para identificar valores em vermelho
            for row in ws.iter_rows(min_row=2, max_col=3, max_row=ws.max_row):
                cell = row[2]  # Coluna de VALOR
                if cell.font.color and cell.font.color.rgb == 'FFFF0000':  # Vermelho
                    df.at[cell.row - 1, 'VALOR'] = f"-{cell.value}"  # Adiciona um "-" para indicar pagamento
        else:
            raise ValueError("Formato de arquivo não suportado. Use CSV ou XLSX.")

        # Verifica se o DataFrame tem pelo menos 3 colunas
        if df.shape[1] < 3:
            raise ValueError("A planilha deve ter pelo menos 3 colunas (DATA, DESCRICAO, VALOR).")

        # Renomeia as colunas para garantir consistência
        df.columns = ["DATA", "DESCRICAO", "VALOR"] + list(df.columns[3:])

        # Converte a coluna VALOR para numérico, tratando valores negativos
        df["VALOR"] = pd.to_numeric(df["VALOR"], errors="coerce")

        return df

    except Exception as e:
        raise Exception(f"Erro ao carregar a planilha: {e}")


def conciliar_pagos_banco(planilha_banco, planilha_pagos):
    """Concilia as planilhas de banco e pagos, gerando uma nova planilha."""
    try:
        # Carrega as planilhas
        banco_df = carregar_planilha(planilha_banco)
        pagos_df = carregar_planilha(planilha_pagos)

        # Verifica se as planilhas têm as colunas necessárias
        if banco_df.shape[1] < 3 or pagos_df.shape[1] < 3:
            return {"status": "erro", "message": "As planilhas devem ter pelo menos 3 colunas (DATA, DESCRICAO, VALOR)."}

        # Converte as colunas DATA para o mesmo tipo (datetime)
        banco_df["DATA"] = pd.to_datetime(banco_df["DATA"], format="%d/%m/%Y", errors="coerce")
        pagos_df["DATA"] = pd.to_datetime(pagos_df["DATA"], format="%d/%m/%Y", errors="coerce")

        # Converte as colunas VALOR para o mesmo tipo (float)
        banco_df["VALOR"] = pd.to_numeric(banco_df["VALOR"], errors="coerce")
        pagos_df["VALOR"] = pd.to_numeric(pagos_df["VALOR"], errors="coerce")

        # Lista para armazenar os resultados
        resultado = []

        # Itera sobre as linhas do banco
        for i, linha_banco in banco_df.iterrows():
            data_banco = linha_banco["DATA"]
            valor_banco = linha_banco["VALOR"]

            # Procura correspondência na planilha de pagos
            correspondencia = pagos_df[
                (pagos_df["DATA"] == data_banco) & (
                    pagos_df["VALOR"] == valor_banco)
            ]

            if not correspondencia.empty:
                # Substitui a descrição do banco pela descrição de pagos
                linha_banco["DESCRICAO"] = correspondencia.iloc[0]["DESCRICAO"]
                # Remove a linha de pagos para evitar duplicações
                pagos_df.drop(correspondencia.index, inplace=True)

            # Adiciona a linha do banco ao resultado
            resultado.append(linha_banco)

        # Converte o resultado de volta para um DataFrame
        resultado_df = pd.DataFrame(resultado)

        # Filtra as linhas do banco que não foram conciliadas
        linhas_nao_conciliadas = banco_df[
            ~banco_df.apply(lambda row: (row["DATA"], row["VALOR"]), axis=1).isin(
                resultado_df[["DATA", "VALOR"]].apply(tuple, axis=1)
            )
        ]

        # Adiciona as linhas não conciliadas ao resultado, mantendo a ordem original
        for i, linha in banco_df.iterrows():
            if i in linhas_nao_conciliadas.index:
                resultado_df = pd.concat(
                    [resultado_df, linha.to_frame().T], ignore_index=True)

        # Salva o resultado em um arquivo Excel
        caminho_arquivo = salvar_dados("banco_conciliado.xlsx")
        if caminho_arquivo:
            resultado_df.to_excel(caminho_arquivo, index=False)  # Salva como Excel
            return {"status": "success", "message": "Planilha conciliada salva com sucesso!"}
        else:
            return {"status": "erro", "message": "Nenhum arquivo foi selecionado para salvar."}

    except Exception as e:
        return {"status": "erro", "message": f"Erro ao processar as planilhas: {str(e)}"}


def obter_conciliacao(empresa):
    try:
        # Referência ao documento da empresa
        empresa_ref = db.collection('empresas').document(str(empresa))
        empresa_doc = empresa_ref.get()

        if not empresa_doc.exists:
            return {"success": False, "message": "Empresa não encontrada."}

        # Referência à coleção "conciliacoes" da empresa
        conciliacoes_ref = empresa_ref.collection('conciliacoes')
        conciliacoes = conciliacoes_ref.stream()

        # Lista para armazenar as conciliações
        conciliacoes_lista = []

        # Itera sobre os documentos da coleção "conciliacoes"
        for conciliacao in conciliacoes:
            conciliacao_data = conciliacao.to_dict()
            # Adiciona o ID do documento
            conciliacao_data['id'] = conciliacao.id
            conciliacoes_lista.append(conciliacao_data)

        if not conciliacoes_lista:
            return {"success": True, "message": "Nenhuma conciliação encontrada.", "conciliacoes": []}

        return {"success": True, "conciliacoes": conciliacoes_lista}

    except Exception as e:
        return {"success": False, "message": f"Erro ao obter conciliações: {e}"}


def conciliar_pagos_banco_conta(planilha_banco, planilha_pagos, numeroEmpresa, numeroBanco):
    """Concilia as planilhas de banco e pagos, gerando uma nova planilha com filtros adicionais."""
    try:
        # Carrega as planilhas
        banco_df = carregar_planilha(planilha_banco)
        pagos_df = carregar_planilha(planilha_pagos)

        # Verifica se as planilhas têm as colunas necessárias
        if banco_df.shape[1] < 3 or pagos_df.shape[1] < 3:
            return {"status": "erro", "message": "As planilhas devem ter pelo menos 3 colunas (DATA, DESCRICAO, VALOR)."}

        # Converte as colunas DATA para o mesmo tipo (datetime)
        banco_df["DATA"] = pd.to_datetime(banco_df["DATA"], format="%d/%m/%Y", errors="coerce")
        pagos_df["DATA"] = pd.to_datetime(pagos_df["DATA"], format="%d/%m/%Y", errors="coerce")

        # Converte as colunas VALOR para o mesmo tipo (float)
        banco_df["VALOR"] = pd.to_numeric(banco_df["VALOR"], errors="coerce")
        pagos_df["VALOR"] = pd.to_numeric(pagos_df["VALOR"], errors="coerce")

        # Obtém as conciliações da empresa
        conciliacoes = obter_conciliacao(numeroEmpresa)
        if not conciliacoes["success"]:
            return {"status": "erro", "message": conciliacoes["message"]}

        # Filtra as conciliações pelo número do banco
        conciliacoes_filtradas = [
            conc for conc in conciliacoes["conciliacoes"] if conc.get("banco") == numeroBanco
        ]

        # Itera sobre as linhas do banco
        for i, linha_banco in banco_df.iterrows():
            data_banco = linha_banco["DATA"]
            valor_banco = linha_banco["VALOR"]

            # Procura correspondência na planilha de pagos
            correspondencia = pagos_df[
                (pagos_df["DATA"] == data_banco) & (pagos_df["VALOR"] == valor_banco)
            ]

            if not correspondencia.empty:
                # Substitui a descrição do banco pela descrição de pagos
                banco_df.at[i, "DESCRICAO"] = correspondencia.iloc[0]["DESCRICAO"]
                # Remove a linha de pagos para evitar duplicações
                pagos_df.drop(correspondencia.index, inplace=True)

        # Adiciona as colunas DEBITO e CREDITO ao DataFrame, se ainda não existirem
        if "DEBITO" not in banco_df.columns:
            banco_df["DEBITO"] = ""  # Inicializa com valores vazios
        if "CREDITO" not in banco_df.columns:
            banco_df["CREDITO"] = ""  # Inicializa com valores vazios

        # Atualiza a planilha do banco com as informações das conciliações
        for conc in conciliacoes_filtradas:
            descricao_firebase = conc.get("descricao", "")
            debito_firebase = conc.get("debito", "")
            credito_firebase = conc.get("credito", "")

            # Verifica se a descrição do Firebase está contida na descrição do banco
            banco_df.loc[
                banco_df["DESCRICAO"].str.contains(descricao_firebase, case=False, na=False, regex=False),
                "DEBITO"
            ] = int(float(debito_firebase))

            banco_df.loc[
                banco_df["DESCRICAO"].str.contains(descricao_firebase, case=False, na=False, regex=False),
                "CREDITO"
            ] = int(float(credito_firebase))

        # Substitui NaN por string vazia nas colunas DEBITO e CREDITO
        banco_df["DEBITO"] = banco_df["DEBITO"].fillna("")
        banco_df["CREDITO"] = banco_df["CREDITO"].fillna("")

        # Garante que as colunas DEBITO e CREDITO sejam tratadas como string (formato geral)
        banco_df["DEBITO"] = banco_df["DEBITO"].astype(str).replace(r'\.0', '', regex=True)
        banco_df["CREDITO"] = banco_df["CREDITO"].astype(str).replace(r'\.0', '', regex=True)

        # Salva o resultado em um arquivo Excel
        caminho_arquivo = salvar_dados("banco_conciliado.xlsx")
        if caminho_arquivo:
            banco_df.to_excel(caminho_arquivo, index=False)  # Salva como Excel
            return {"status": "success", "message": "Planilha conciliada salva com sucesso!"}
        else:
            return {"status": "erro", "message": "Nenhum arquivo foi selecionado para salvar."}

    except Exception as e:
        return {"status": "erro", "message": f"Erro ao processar as planilhas: {str(e)}"}
    
def conciliar_apenas_banco(planilha_banco, numeroEmpresa, numeroBanco):
    """Concilia a planilha do banco com as informações do Firebase, sem a planilha de pagos."""
    try:
        # Carrega a planilha do banco
        banco_df = carregar_planilha(planilha_banco)

        # Renomeia dinamicamente as três primeiras colunas
        banco_df.columns = ["DATA", "DESCRICAO", "VALOR"] + list(banco_df.columns[3:])

        # Converte as colunas para os tipos corretos
        banco_df["DATA"] = pd.to_datetime(banco_df["DATA"], format="%d/%m/%Y", errors="coerce")
        banco_df["VALOR"] = banco_df["VALOR"].astype(str).replace(r'\.0$', '', regex=True)

        # Obtém as conciliações da empresa
        conciliacoes = obter_conciliacao(numeroEmpresa)
        if not conciliacoes["success"]:
            return {"status": "erro", "message": conciliacoes["message"]}

        # Filtra as conciliações pelo número do banco
        conciliacoes_filtradas = [
            conc for conc in conciliacoes["conciliacoes"] if conc.get("banco") == numeroBanco
        ]

        # Aplica as conciliações na planilha do banco
        for conc in conciliacoes_filtradas:
            descricao_firebase = conc.get("descricao", "")
            debito_firebase = conc.get("debito", "")

            # Verifica se a descrição do Firebase está contida na descrição do banco
            banco_df.loc[
                banco_df["DESCRICAO"].str.contains(descricao_firebase, case=False, na=False),
                "DEBITO"
            ] = int(float(debito_firebase))

        # Substitui NaN por string vazia na coluna DEBITO
        banco_df["DEBITO"] = banco_df["DEBITO"].fillna("")

        # Garante que a coluna DEBITO seja tratada como string (formato geral)
        banco_df["DEBITO"] = banco_df["DEBITO"].astype(str).replace(r'\.0', '', regex=True)

        # Salva o resultado em um arquivo CSV
        caminho_arquivo = salvar_dados("banco_conciliado.csv")
        if caminho_arquivo:
            banco_df.to_csv(caminho_arquivo, index=False, sep=';', encoding='utf-8-sig')
            return {"status": "success", "message": "Planilha conciliada salva com sucesso!"}
        else:
            return {"status": "erro", "message": "Nenhum arquivo foi selecionado para salvar."}

    except Exception as e:
        return {"status": "erro", "message": f"Erro ao processar a planilha: {str(e)}"}



def main():
    try:
        if len(sys.argv) < 2:
            raise ValueError("Argumentos insuficientes.")

        action = sys.argv[1]

        if action == 'conciliar_pagos_banco':
            if len(sys.argv) < 4:
                raise ValueError("Caminho do banco e pagos não fornecidos.")
            caminho_banco = sys.argv[2]
            caminho_pagos = sys.argv[3]
            result = conciliar_pagos_banco(caminho_banco, caminho_pagos)

        elif action == 'conciliar_pagos_banco_conta':
            if len(sys.argv) < 6:
                raise ValueError("Argumentos insuficientes para conciliar com conta.")
            caminho_banco = sys.argv[2]
            caminho_pagos = sys.argv[3]
            numeroEmpresa = int(sys.argv[4])
            numeroBanco = int(sys.argv[5])
            result = conciliar_pagos_banco_conta(caminho_banco, caminho_pagos, numeroEmpresa, numeroBanco)
            
        elif action == 'conciliar_apenas_banco':
            if len(sys.argv) < 5:
                raise ValueError("Argumentos insuficientes para conciliar apenas o banco.")
            caminho_banco = sys.argv[2]
            numeroEmpresa = int(sys.argv[3])
            numeroBanco = int(sys.argv[4])
            result = conciliar_apenas_banco(caminho_banco, numeroEmpresa, numeroBanco)

        else:
            raise ValueError("Ação desconhecida.")

        print(json.dumps(result))

    except Exception as e:
        print(json.dumps({"status": "erro", "message": str(e)}))

if __name__ == '__main__':
    main()
